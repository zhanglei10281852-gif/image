#!/usr/bin/env python3
from __future__ import annotations

import csv
import argparse
from datetime import datetime, timezone, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from PIL import Image as PILImage


ROOT = Path(__file__).resolve().parent
DEFAULT_MANIFEST = ROOT / "data" / "desktop_scenes" / "manifest.csv"
FALLBACK_MANIFEST = ROOT / "sample_run" / "manifest.csv"
DEFAULT_OUTPUT = ROOT / "outputs" / "桌面场景图片标注表_张磊.xlsx"
AUTHOR = "张磊"
BASE_HEADERS = ["自动编号", "来源URL", "图片URL", "场景环境", "创建时间", "做题人", "重复标记"]
EXTRA_HEADERS = [
    "record_id",
    "source_type",
    "source_platform",
    "normalized_source_url",
    "fetched_at",
    "query",
    "title",
    "author",
    "license_type",
    "license_url",
    "image_path",
    "width",
    "height",
    "short_edge",
    "scene_setting",
    "complexity_level",
    "risk_flag",
    "notes",
]
HEADERS = BASE_HEADERS
SHANGHAI = timezone(timedelta(hours=8))
SCENE_LABELS = {
    "indoor": "室内",
    "outdoor": "室外",
    "semi_outdoor": "半户外",
    "室内": "室内",
    "室外": "室外",
    "半户外": "半户外",
}


def parse_time(value: str) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).astimezone(SHANGHAI).replace(tzinfo=None)
    except ValueError:
        return None


def infer_scene_cn(row: dict[str, str]) -> str:
    scene = (row.get("scene_setting") or "").strip()
    text = " ".join(
        [
            row.get("query", ""),
            row.get("title", ""),
            row.get("source_url", ""),
            row.get("notes", ""),
        ]
    ).lower()

    outdoor_words = ["室外", "户外", "野餐", "露营", "picnic", "outdoor", "camping", "garden", "park", "beach", "street table", "outside"]
    semi_words = ["半户外", "阳台", "露台", "外摆", "balcony", "patio", "terrace", "porch", "veranda", "semi-outdoor", "sidewalk cafe", "outdoor seating", "al fresco"]
    indoor_words = ["home", "indoor", "office", "desk", "workspace", "study", "kitchen", "living room", "dining room", "bedroom", "bathroom", "counter", "countertop", "workbench", "bench", "vanity", "nightstand", "bedside", "cafe", "café", "coffee shop", "restaurant", "table", "laptop", "notebook", "book", "breakfast", "tea", "coffee", "cup", "plate", "lamp", "makeup", "cosmetics", "meeting", "craft", "tools", "art", "sewing", "书桌", "办公", "工位", "厨房", "室内", "台面", "餐桌", "咖啡桌", "茶几", "工作台"]

    if scene == "semi_outdoor" and not any(word in text for word in semi_words + outdoor_words):
        scene = ""
    if scene in SCENE_LABELS:
        return SCENE_LABELS[scene]

    if any(word in text for word in outdoor_words):
        return "室外"
    if any(word in text for word in semi_words):
        return "半户外"
    if any(word in text for word in indoor_words):
        return "室内"
    return ""


def infer_complexity_level(row: dict[str, str]) -> str:
    current = (row.get("complexity_level") or "").strip()
    if current in {"L1", "L2", "L3"}:
        return current

    text = " ".join(
        [
            row.get("query", ""),
            row.get("title", ""),
            row.get("source_url", ""),
            row.get("notes", ""),
        ]
    ).lower()
    l3_words = [
        "workbench",
        "workshop",
        "tools",
        "parts",
        "craft",
        "materials",
        "meeting",
        "kitchen countertop",
        "ingredients",
        "picnic",
        "outdoor dining",
        "restaurant",
        "breakfast table",
        "dining table",
        "多物体",
        "工具",
        "食材",
        "餐桌",
        "野餐",
    ]
    l1_words = [
        "single",
        "one cup",
        "coffee cup",
        "mug",
        "tea cup",
        "cup on",
        "plate on",
        "book on",
        "minimal",
        "closeup",
        "close-up",
        "杯子",
        "单个",
        "近景",
    ]
    if any(word in text for word in l3_words):
        return "L3"
    if any(word in text for word in l1_words):
        return "L1"
    return "L2"


def is_strict_valid(row: dict[str, str]) -> bool:
    if row.get("risk_flag"):
        return False
    if (row.get("license_type") or "").strip() == "restricted":
        return False
    if not row.get("source_url") or not row.get("source_image_url"):
        return False
    if "/thumb/" in (row.get("source_image_url") or "").lower():
        return False
    try:
        if int(row.get("short_edge") or 0) < 512:
            return False
    except ValueError:
        return False
    return True


def read_records(
    manifest: Path,
    include_risky: bool = False,
    offset: int = 0,
    limit: int | None = None,
) -> list[dict[str, str]]:
    if not manifest.exists():
        return []
    records: list[dict[str, str]] = []
    seen_sources: set[str] = set()
    with manifest.open("r", newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            if not include_risky and not is_strict_valid(row):
                continue
            normalized = (row.get("normalized_source_url") or row.get("source_url") or "").strip()
            if not normalized or normalized in seen_sources or row.get("risk_flag") == "duplicate_source_url":
                continue
            seen_sources.add(normalized)
            records.append(row)

    selected = records[max(offset, 0) :]
    if limit is not None and limit >= 0:
        selected = selected[:limit]
    return selected


def read_rows(
    manifest: Path,
    include_risky: bool = False,
    include_extra_fields: bool = False,
    include_thumbnails: bool = False,
    offset: int = 0,
    limit: int | None = None,
) -> list[list[object]]:
    selected = read_records(manifest, include_risky=include_risky, offset=offset, limit=limit)
    rows: list[list[object]] = []
    for row in selected:
            created_at = parse_time(row.get("fetched_at", ""))
            scene_cn = infer_scene_cn(row)
            complexity_level = infer_complexity_level(row)
            base_values: list[object] = [
                offset + len(rows) + 1,
                row.get("source_url", ""),
                row.get("source_image_url", ""),
                scene_cn,
                created_at,
                AUTHOR,
                "不重复",
            ]
            if include_thumbnails:
                base_values.insert(1, "")
            if include_extra_fields:
                extra_values = [
                    row.get("record_id", ""),
                    row.get("source_type", ""),
                    row.get("source_platform", ""),
                    row.get("normalized_source_url", ""),
                    row.get("fetched_at", ""),
                    row.get("query", ""),
                    row.get("title", ""),
                    row.get("author", ""),
                    row.get("license_type", ""),
                    row.get("license_url", ""),
                    row.get("image_path", ""),
                    parse_int(row.get("width")),
                    parse_int(row.get("height")),
                    parse_int(row.get("short_edge")),
                    scene_cn_to_manifest(scene_cn),
                    complexity_level,
                    row.get("risk_flag", ""),
                    row.get("notes", ""),
                ]
                base_values.extend(extra_values)
            rows.append(base_values)
    return rows


def parse_int(value: str | None) -> int | str:
    if value in (None, ""):
        return ""
    try:
        return int(value)
    except ValueError:
        return value


def scene_cn_to_manifest(value: str) -> str:
    return {"室内": "indoor", "室外": "outdoor", "半户外": "semi_outdoor"}.get(value, "")


def build_headers(include_extra_fields: bool, include_thumbnails: bool) -> list[str]:
    headers = BASE_HEADERS.copy()
    if include_thumbnails:
        headers.insert(1, "缩略图")
    if include_extra_fields:
        headers.extend(EXTRA_HEADERS)
    return headers


def create_thumbnail(source: Path, dest: Path, max_size: tuple[int, int] = (120, 86)) -> Path | None:
    if not source.exists():
        return None
    try:
        with PILImage.open(source) as im:
            im = im.convert("RGB")
            im.thumbnail(max_size)
            dest.parent.mkdir(parents=True, exist_ok=True)
            im.save(dest, "JPEG", quality=82, optimize=True)
        return dest
    except Exception:
        return None


def add_thumbnails(ws, manifest: Path, records: list[dict[str, str]], output: Path) -> None:
    thumb_dir = output.parent / f".{output.stem}_thumbs"
    ws.column_dimensions["B"].width = 18
    for idx, row in enumerate(records, start=2):
        image_path = row.get("image_path", "")
        if not image_path:
            continue
        source = manifest.parent / image_path
        thumb = create_thumbnail(source, thumb_dir / f"{row.get('record_id') or idx}.jpg")
        if not thumb:
            continue
        image = XLImage(str(thumb))
        image.anchor = f"B{idx}"
        ws.add_image(image)
        ws.row_dimensions[idx].height = 70


def build_workbook(
    manifest: Path,
    output: Path,
    include_risky: bool = False,
    include_extra_fields: bool = False,
    include_thumbnails: bool = False,
    offset: int = 0,
    limit: int | None = None,
) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "桌面场景标注"
    ws.sheet_view.showGridLines = False

    headers = build_headers(include_extra_fields=include_extra_fields, include_thumbnails=include_thumbnails)
    records = read_records(manifest, include_risky=include_risky, offset=offset, limit=limit)
    data_rows = read_rows(
        manifest,
        include_risky=include_risky,
        include_extra_fields=include_extra_fields,
        include_thumbnails=include_thumbnails,
        offset=offset,
        limit=limit,
    )

    ws.append(headers)
    for row in data_rows:
        ws.append(row)

    used_rows = max(len(data_rows) + 1, 2)
    validation_rows = 100000
    header_fill = PatternFill("solid", fgColor="F6F7F9")
    body_fill = PatternFill("solid", fgColor="FFFFFF")
    border_side = Side(style="thin", color="DADDE1")
    grid_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    header_font = Font(name="Arial", size=13, bold=True, color="1F2328")
    body_font = Font(name="Arial", size=12, color="1F2328")
    link_font = Font(name="Arial", size=12, color="0969DA", underline="single")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = grid_border

    for row in ws.iter_rows(min_row=2, max_row=used_rows, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.fill = body_fill
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = grid_border

    for row_idx in range(2, len(data_rows) + 2):
        source_col = 3 if include_thumbnails else 2
        image_col = 4 if include_thumbnails else 3
        created_col = 6 if include_thumbnails else 5
        hyperlink_cols = [source_col, image_col]
        if include_extra_fields:
            hyperlink_cols.append(18 if include_thumbnails else 16)
        for col_idx in hyperlink_cols:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.hyperlink = cell.value
                cell.font = link_font
        ws.cell(row=row_idx, column=created_col).number_format = "yyyy-mm-dd hh:mm"

    scene_validation = DataValidation(type="list", formula1='"室内,室外,半户外"', allow_blank=True)
    scene_validation.error = "请选择：室内、室外、半户外"
    scene_validation.errorTitle = "无效场景环境"
    scene_validation.prompt = "请选择场景环境"
    scene_validation.promptTitle = "场景环境"
    ws.add_data_validation(scene_validation)
    scene_col_letter = "E" if include_thumbnails else "D"
    scene_validation.add(f"{scene_col_letter}2:{scene_col_letter}{validation_rows}")

    duplicate_validation = DataValidation(type="list", formula1='"不重复,重复"', allow_blank=False)
    duplicate_validation.error = "请选择：不重复 或 重复"
    duplicate_validation.errorTitle = "无效重复标记"
    duplicate_validation.prompt = "请选择重复标记"
    duplicate_validation.promptTitle = "重复标记"
    ws.add_data_validation(duplicate_validation)
    duplicate_col_letter = "H" if include_thumbnails else "G"
    duplicate_validation.add(f"{duplicate_col_letter}2:{duplicate_col_letter}{validation_rows}")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).column_letter}{max(len(data_rows) + 1, 2)}"

    if include_thumbnails:
        widths = {
            "A": 14,
            "B": 18,
            "C": 46,
            "D": 54,
            "E": 16,
            "F": 22,
            "G": 14,
            "H": 16,
        }
    else:
        widths = {
        "A": 14,
        "B": 46,
        "C": 54,
        "D": 16,
        "E": 22,
        "F": 14,
        "G": 16,
        }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    if include_extra_fields:
        extra_widths = [22, 14, 20, 44, 24, 28, 44, 22, 18, 44, 26, 12, 12, 12, 16, 18, 18, 36]
        start_col = len(BASE_HEADERS) + (1 if include_thumbnails else 0) + 1
        for index, width in enumerate(extra_widths, start=start_col):
            ws.column_dimensions[get_column_letter(index)].width = width
    ws.row_dimensions[1].height = 36
    for row_idx in range(2, used_rows + 1):
        if not include_thumbnails:
            ws.row_dimensions[row_idx].height = 30

    ws[f"{scene_col_letter}1"].comment = Comment("下拉选项：室内、室外、半户外", AUTHOR)
    ws[f"{duplicate_col_letter}1"].comment = Comment("下拉选项：不重复、重复", AUTHOR)

    if include_thumbnails:
        add_thumbnails(ws, manifest, records, output)

    wb.save(output)


def verify_workbook(output: Path) -> None:
    wb = load_workbook(output)
    ws = wb["桌面场景标注"]
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    has_thumbnails = len(headers) > 1 and headers[1] == "缩略图"
    expected_base = BASE_HEADERS.copy()
    if has_thumbnails:
        expected_base.insert(1, "缩略图")
    if headers[: len(expected_base)] != expected_base:
        raise RuntimeError(f"Header mismatch: {headers}")
    author_cell = "G2" if has_thumbnails else "F2"
    if ws[author_cell].value not in (AUTHOR, None):
        raise RuntimeError("Author column was not populated as expected")
    validations = list(ws.data_validations.dataValidation)
    if len(validations) < 2:
        raise RuntimeError("Expected scene and duplicate data validations")
    print(f"created={output}")
    print(f"rows={ws.max_row - 1}")
    print(f"headers={headers}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build tabletop scene labeling Excel.")
    parser.add_argument("--manifest", default=str(DEFAULT_MANIFEST if DEFAULT_MANIFEST.exists() else FALLBACK_MANIFEST))
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--include-risky", action="store_true", help="include rows with risk flags; off by default for strict delivery")
    parser.add_argument("--include-extra-fields", action="store_true", help="append manifest/spec fields after the original 7 columns")
    parser.add_argument("--include-thumbnails", action="store_true", help="insert a visible thumbnail column before source URL")
    parser.add_argument("--offset", type=int, default=0, help="zero-based offset after strict filtering and dedupe")
    parser.add_argument("--limit", type=int, default=-1, help="maximum rows to export after offset; negative means all")
    args = parser.parse_args()
    output_path = Path(args.output)
    build_workbook(
        Path(args.manifest),
        output_path,
        include_risky=args.include_risky,
        include_extra_fields=args.include_extra_fields,
        include_thumbnails=args.include_thumbnails,
        offset=args.offset,
        limit=None if args.limit < 0 else args.limit,
    )
    verify_workbook(output_path)
